from random import randint

from openpyxl import Workbook

wb = Workbook()
ws1 = wb.active
ws1.title = "NadoSheet"

# A1 셀에 1 이라는 값을 입력
ws1["A1"] = 1
ws1["A2"] = 2
ws1["A3"] = 3

ws1["B1"] = 4
ws1["B2"] = 5
ws1["B3"] = 6

print(ws1["A1"])  # A1 셀의 정보를 출력
print(ws1["A1"].value)  # A1 셀의 '값'을 출력
print(ws1["A10"].value)  # 값이 업을 때는 'None' 을 출력

# row = 1, 2, 3, ...
# column = A(1), B(2), C(3), ...
print(ws1.cell(column=2, row=1).value)  # ws1["B1"].value
print(ws1.cell(column=1, row=1).value)  # ws1["A1"].value

c = ws1.cell(column=3, row=1, value=10)  # ws["C1"] = 10, ws["C1"].value = 10
print(c.value)  # ws1["C1"].value

# 반복문을 이용해서 랜덤 숫자 채우기
index = 1
for x in range(1, 11):  # 10 개 row
    for y in range(1, 11):  # 10 개 row
        ws1.cell(row=x, column=y, value=randint(0, 100))  # 0~100 사이의 숫자
        # ws1.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample.xlsx")
